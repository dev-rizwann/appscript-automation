# web.py
from pathlib import Path
from datetime import datetime, date
import re

import pdfplumber
import pandas as pd
from openpyxl.utils import get_column_letter

# =====================
# Regex definitions
# =====================
DATE_RE = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
TXN_RE = re.compile(r"^\d{4}$")
ORDER_LONG_RE = re.compile(r"^\d{10,}(?:_\d+)?$")
NUM_RE = re.compile(r"^\d+(?:\.\d{1,2})?$")

TOTAL_USD_RE = re.compile(
    r"Total\s*\(\s*USD\s*\)\s*:\s*([$]?\s*[\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)
TOTAL_AMOUNT_RE = re.compile(
    r"TOTAL\s+AMOUNT\s*[$]?\s*([\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)

SINGLE_COUNTRIES = {
    "canada", "australia", "mexico", "china", "france", "germany",
    "italy", "spain", "japan", "singapore", "uae", "pakistan", "uk"
}

# =====================
# Helper functions
# =====================
def clean(t: str) -> str:
    return (t or "").strip().strip("[]{}(),;:")

def parse_ymd_to_date(s: str) -> date:
    y, m, d = s.split("/")
    return date(int(y), int(m), int(d))

def excel_serial_from_date(dt: date) -> int:
    excel_epoch = date(1899, 12, 30)
    return (dt - excel_epoch).days

def extract_tokens_and_text(pdf_path: Path):
    tokens = []
    full_text_parts = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if txt:
                full_text_parts.append(txt)
                tokens.extend(re.split(r"\s+", txt))

    tokens = [t for t in tokens if t]  # remove empty tokens
    return tokens, "\n".join(full_text_parts)

def extract_total_usd(full_text: str):
    if not full_text:
        return None

    m = TOTAL_USD_RE.search(full_text) or TOTAL_AMOUNT_RE.search(full_text)
    if not m:
        return None

    raw = (m.group(1) or "").replace("$", "").replace(" ", "").replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return None

def is_country_at(tokens, i):
    t = clean(tokens[i]).lower()

    if t == "united" and i + 1 < len(tokens):
        if clean(tokens[i + 1]).lower() == "states":
            return True, i + 1

    if t in SINGLE_COUNTRIES:
        return True, i

    return False, -1

def is_row_start(tokens, i):
    if i + 2 >= len(tokens):
        return False

    return (
        ORDER_LONG_RE.match(clean(tokens[i]))
        and TXN_RE.match(clean(tokens[i + 1]))
        and DATE_RE.match(clean(tokens[i + 2]))
    )

def qty_before_country(tokens, row_start, country_pos):
    for j in range(country_pos - 1, row_start, -1):
        t = clean(tokens[j])
        if t.isdigit():
            v = int(t)
            if 1 <= v <= 999:
                return v
    return None

def looks_like_price(tok: str):
    t = clean(tok)

    if TXN_RE.match(t):
        return False
    if not NUM_RE.match(t):
        return False
    if ORDER_LONG_RE.match(t):
        return False

    try:
        v = float(t)
    except ValueError:
        return False

    if "." not in t and v < 10:
        return False
    if v <= 0 or v > 100000:
        return False

    return True

def last_price_after_country(tokens, start, row_end):
    last = None
    for j in range(start, row_end):
        if looks_like_price(tokens[j]):
            last = float(clean(tokens[j]))
    return last

def stop_at_total_usd(tokens):
    for i, tok in enumerate(tokens):
        if clean(tok).lower().startswith("total(usd"):
            return tokens[:i]
    return tokens

def to_dutch_text(num):
    if num is None:
        return ""
    return f"{float(num):.2f}".replace(".", ",")

# =====================
# Core parsing logic
# =====================
def parse_pdf_tokens(tokens, file_name):
    tokens = stop_at_total_usd(tokens)
    rows = []
    i = 0
    n = len(tokens)

    while i < n:
        if not is_row_start(tokens, i):
            i += 1
            continue

        txn_s = clean(tokens[i + 1])
        date_s = clean(tokens[i + 2])

        order4 = int(txn_s) if TXN_RE.match(txn_s) else None
        dt = parse_ymd_to_date(date_s) if DATE_RE.match(date_s) else None
        date_serial = excel_serial_from_date(dt) if dt else None

        row_end = min(n, i + 320)
        j = i + 3
        while j < row_end:
            if is_row_start(tokens, j):
                row_end = j
                break
            j += 1

        country_pos = -1
        country_end = -1
        k = i + 3
        while k < row_end:
            ok, cend = is_country_at(tokens, k)
            if ok:
                country_pos = k
                country_end = cend
                break
            k += 1

        qty = None
        cost = None

        if country_pos != -1:
            qty = qty_before_country(tokens, i, country_pos)
            cost = last_price_after_country(tokens, country_end + 1, row_end)

        if order4 and date_serial and cost is not None:
            rows.append({
                "File Name": file_name,
                "DateSerial": date_serial,
                "Order #": order4,
                "Qty": int(qty) if qty else None,
                "Cost": float(cost),
                "Cost_NL": to_dutch_text(cost),
            })

        i = row_end

    return rows

# =====================
# JSON output for Sheets (3 tabs)
# =====================
def convert_pdfs_to_json(pdf_paths):
    all_rows = []
    totals_rows = []
    log_rows = []

    for pdf in pdf_paths:
        pdf_path = Path(pdf)
        fname = pdf_path.name

        try:
            tokens, full_text = extract_tokens_and_text(pdf_path)
            rows = parse_pdf_tokens(tokens, fname)
            all_rows.extend(rows)

            total_usd = extract_total_usd(full_text)
            totals_rows.append({
                "File Name": fname,
                "Total_USD_Extracted": total_usd,
            })

            log_rows.append({
                "File": fname,
                "Tokens": len(tokens),
                "Rows": len(rows),
                "Status": "OK",
                "Error": "",
            })

        except Exception as e:
            totals_rows.append({
                "File Name": fname,
                "Total_USD_Extracted": None,
            })
            log_rows.append({
                "File": fname,
                "Tokens": "",
                "Rows": "",
                "Status": "ERROR",
                "Error": str(e),
            })

    # Sum COGS by file
    cogs_sum_by_file = {}
    for r in all_rows:
        fn = r.get("File Name")
        cogs_sum_by_file[fn] = cogs_sum_by_file.get(fn, 0) + float(r.get("Cost") or 0)

    # Add computed totals fields
    for t in totals_rows:
        fn = t.get("File Name")
        total = t.get("Total_USD_Extracted")
        cogs_sum = float(cogs_sum_by_file.get(fn, 0))

        t["COGS_Sum"] = cogs_sum
        if total is None:
            t["Diff"] = None
            t["Match"] = "CHECK"
        else:
            diff = cogs_sum - float(total)
            t["Diff"] = diff
            t["Match"] = "OK" if abs(diff) < 0.01 else "CHECK"

    return {
        "status": "success",
        "COGS": {
            "columns": ["File Name", "DateSerial", "Order #", "Qty", "Cost", "Cost_NL"],
            "rows": [
                [
                    r.get("File Name"),
                    r.get("DateSerial"),
                    r.get("Order #"),
                    r.get("Qty"),
                    r.get("Cost"),
                    r.get("Cost_NL"),
                ]
                for r in all_rows
            ],
        },
        "InvoiceTotals": {
            "columns": ["File Name", "Total_USD_Extracted", "COGS_Sum", "Diff", "Match"],
            "rows": [
                [
                    t.get("File Name"),
                    t.get("Total_USD_Extracted"),
                    t.get("COGS_Sum"),
                    t.get("Diff"),
                    t.get("Match"),
                ]
                for t in totals_rows
            ],
        },
        "Log": {
            "columns": ["File", "Tokens", "Rows", "Status", "Error"],
            "rows": [
                [
                    l.get("File"),
                    l.get("Tokens"),
                    l.get("Rows"),
                    l.get("Status"),
                    l.get("Error"),
                ]
                for l in log_rows
            ],
        },
    }

# =====================
# Optional: Keep your old Excel generator too (if you still want)
# =====================
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def apply_formats(writer, sheet_name="COGS"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    for r in range(2, last_row + 1):
        if "DateSerial" in headers:
            ws.cell(r, headers["DateSerial"]).number_format = "0"
        if "Order #" in headers:
            ws.cell(r, headers["Order #"]).number_format = "0"
        if "Qty" in headers:
            ws.cell(r, headers["Qty"]).number_format = "0"
        if "Cost" in headers:
            ws.cell(r, headers["Cost"]).number_format = "#,##0.00"
        if "Cost_NL" in headers:
            ws.cell(r, headers["Cost_NL"]).number_format = "@"

    autosize_columns(ws)

def apply_formats_totals(writer, sheet_name="InvoiceTotals"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    for r in range(2, last_row + 1):
        if "Total_USD_Extracted" in headers:
            ws.cell(r, headers["Total_USD_Extracted"]).number_format = "#,##0.00"

    autosize_columns(ws)
